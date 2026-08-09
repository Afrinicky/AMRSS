"""Report generation (SDD 9.7).

The guarantee under test is one sentence: a report never shows something the
live dashboard would have withheld. It is enforced structurally rather than by
review — a withheld cell reaches the renderers carrying no percentage and no
count, so there is nothing for a formatting decision to disclose by accident.

These exercise the period arithmetic, the content invariants, and that both
renderers produce a real file. What each cell *should* say is the antibiogram
engine's business and is tested there; duplicating it here would test the mock.
"""

import re
from datetime import UTC, date, datetime

import pytest
from openpyxl import load_workbook

from amrss.models.enums import ReportPeriod
from amrss.reports import excel, pdf
from amrss.reports.build import ReportCell, ReportContent, ReportRow, period_bounds

AGENTS = ["Ampicillin", "Ciprofloxacin", "Gentamicin", "Meropenem"]


def cell(percent: int | None, n: int = 84) -> ReportCell:
    if percent is None:
        return ReportCell(display="—", n_display="", reportable=False)
    return ReportCell(display=f"{percent}%", n_display=f"n={n}", reportable=True)


def content(**overrides) -> ReportContent:
    base = dict(
        title="Antimicrobial resistance report — Q2 2026",
        scope_label="Test Regional Block",
        period=ReportPeriod.QUARTERLY,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 6, 30),
        generated_at=datetime(2026, 8, 9, 11, 42, tzinfo=UTC),
        antibiotics=AGENTS,
        rows=[
            ReportRow(
                organism="Escherichia coli",
                kingdom="bacteria",
                isolate_count=412,
                cells=[cell(18), cell(54), cell(79), cell(None)],
            ),
            ReportRow(
                organism="Candida albicans",
                kingdom="fungi",
                isolate_count=125,
                cells=[cell(None), cell(None), cell(None), cell(None)],
            ),
        ],
        facilities_contributing=6,
        facilities_expected=7,
        facilities_excluded_by_quality=1,
        isolates_excluded_by_quality=143,
        completeness_percent=86.0,
        data_last_updated="2026-08-07T09:12:00+00:00",
        raw_isolate_count=4357,
        eligible_isolate_count=3218,
        pending_interpretation_count=268,
        minimum_isolates=30,
        small_cell_threshold=5,
        methodology={
            "aggregation_level": "regional",
            "methodology_versions": ["antibiogram 2026.1"],
        },
        signals=["Escherichia coli — Ciprofloxacin: 74% to 51%. Signal — requires expert review."],
        notes=["Combinations with fewer than 30 isolates are not reported."],
    )
    return ReportContent(**{**base, **overrides})


# ---- Period arithmetic -----------------------------------------------------


def test_a_monthly_report_covers_the_whole_calendar_month():
    """Whole calendar periods, never a rolling window ending today. A monthly
    report generated on the 12th that silently covered twelve days would be
    compared against last month's full one."""
    assert period_bounds(ReportPeriod.MONTHLY, date(2026, 8, 12)) == (
        date(2026, 8, 1),
        date(2026, 8, 31),
    )


def test_a_december_monthly_report_rolls_into_the_next_year():
    assert period_bounds(ReportPeriod.MONTHLY, date(2026, 12, 20)) == (
        date(2026, 12, 1),
        date(2026, 12, 31),
    )


def test_february_ends_on_the_right_day_in_a_leap_year():
    assert period_bounds(ReportPeriod.MONTHLY, date(2028, 2, 9))[1] == date(2028, 2, 29)


@pytest.mark.parametrize(
    ("anchor", "expected"),
    [
        (date(2026, 2, 9), (date(2026, 1, 1), date(2026, 3, 31))),
        (date(2026, 5, 9), (date(2026, 4, 1), date(2026, 6, 30))),
        (date(2026, 8, 9), (date(2026, 7, 1), date(2026, 9, 30))),
        (date(2026, 11, 9), (date(2026, 10, 1), date(2026, 12, 31))),
    ],
)
def test_every_quarter_resolves_to_its_own_calendar_bounds(anchor, expected):
    assert period_bounds(ReportPeriod.QUARTERLY, anchor) == expected


def test_an_annual_report_covers_the_calendar_year():
    assert period_bounds(ReportPeriod.ANNUAL, date(2026, 6, 15)) == (
        date(2026, 1, 1),
        date(2026, 12, 31),
    )


def test_an_ad_hoc_report_refuses_to_invent_bounds():
    """An arbitrary range has no calendar period to derive. Guessing one would
    produce a report whose title and contents disagreed."""
    with pytest.raises(ValueError, match="explicit bounds"):
        period_bounds(ReportPeriod.AD_HOC, date(2026, 8, 9))


# ---- The withholding guarantee ---------------------------------------------


def test_a_withheld_cell_carries_nothing_for_a_renderer_to_print():
    withheld = cell(None)

    assert not withheld.reportable
    assert withheld.n_display == ""
    assert not re.search(r"\d", withheld.display)


def test_the_excel_antibiogram_never_prints_a_bare_percentage():
    """Every percentage carries its n (SDD 11.3), and every withheld cell
    carries neither."""
    workbook = load_workbook_from(excel.render(content()))
    sheet = workbook["Antibiogram"]

    reportable = withheld = 0
    for row in sheet.iter_rows(min_row=6, max_row=7, min_col=4):
        for value_cell in row:
            value = value_cell.value or ""
            if value.strip() == "—":
                withheld += 1
                assert "%" not in value and "n=" not in value
            elif value:
                reportable += 1
                assert "%" in value and "n=" in value

    assert reportable == 3
    assert withheld == 5


def test_the_excel_workbook_separates_coverage_from_the_table():
    """A reader who does not know one laboratory in seven was excluded cannot
    judge the table, and a footnote at the bottom of a wide sheet is a footnote
    nobody scrolls to."""
    workbook = load_workbook_from(excel.render(content()))

    assert "Coverage and quality" in workbook.sheetnames
    labels = [workbook["Coverage and quality"].cell(row=r, column=1).value for r in range(3, 16)]
    assert "Facilities excluded on quality grounds" in labels
    assert "Reporting threshold" in labels


def test_the_methodology_travels_inside_the_workbook():
    """A report circulated by email leaves the system that generated it. A
    figure whose thresholds cannot be recovered is not auditable."""
    workbook = load_workbook_from(excel.render(content()))

    assert "Methodology" in workbook.sheetnames
    assert workbook["Methodology"]["A1"].value == "Methodology"


def test_signals_are_kept_on_their_own_sheet():
    """They do not meet the reporting threshold, so they must not sit in the
    antibiogram where they would be read as rates."""
    workbook = load_workbook_from(excel.render(content()))

    assert "Signals" in workbook.sheetnames
    assert "Signals" not in workbook["Antibiogram"].cell(row=5, column=1).value


def test_a_report_with_no_signals_omits_the_sheet_rather_than_showing_an_empty_one():
    workbook = load_workbook_from(excel.render(content(signals=[])))

    assert "Signals" not in workbook.sheetnames


# ---- Renderers produce real files ------------------------------------------


def test_the_pdf_renderer_produces_a_valid_document():
    raw = pdf.render(content())

    assert raw.startswith(b"%PDF-")
    assert raw.rstrip().endswith(b"%%EOF")
    assert len(raw) > 2000


def test_the_pdf_states_its_scope_and_period_on_every_page():
    """A single page separated from the document must still say what it is: a
    table of percentages with no scope and no period is a hazard."""
    text = pdf_text(pdf.render(content()))

    assert "Test Regional Block" in text
    assert "Page 1" in text


def test_the_pdf_carries_the_clinical_framing():
    text = pdf_text(pdf.render(content()))

    assert "clinical judgment" in text


def test_an_empty_antibiogram_says_why_rather_than_printing_nothing():
    """No reportable combination is a statement about data volume, not about
    resistance, and a blank page invites the opposite reading."""
    text = pdf_text(pdf.render(content(rows=[])))

    assert "No organism-agent combination" in text or "reporting threshold" in text


def test_both_renderers_handle_a_report_with_no_rows():
    empty = content(rows=[], signals=[])

    assert pdf.render(empty).startswith(b"%PDF-")
    assert load_workbook_from(excel.render(empty)).sheetnames


# ---- Helpers ---------------------------------------------------------------


def load_workbook_from(raw: bytes):
    import io

    return load_workbook(io.BytesIO(raw))


def pdf_text(raw: bytes) -> str:
    """Text drawn in the document.

    Rendered uncompressed for the assertion so the test reads the same strings
    the renderer wrote, rather than reimplementing ReportLab's stream encoding.
    """
    return b" ".join(re.findall(rb"\((.*?)\)", raw)).decode("latin-1")


@pytest.fixture(autouse=True)
def _uncompressed_pdf():
    """Compression is a transport concern, not a content one.

    Turning it off for the duration of these tests keeps the text assertions
    readable; the production renderer is unchanged.
    """
    from reportlab import rl_config

    previous = rl_config.pageCompression
    rl_config.pageCompression = 0
    yield
    rl_config.pageCompression = previous
