"""Excel rendering of a report.

Formatting only. The renderer receives cells whose values are already resolved
to what may be printed — a withheld cell arrives carrying no percentage and no
count — so there is nothing here for a formatting decision to accidentally
disclose.

The workbook is deliberately several sheets rather than one. A stewardship
committee opens the antibiogram; a data steward opens the methodology sheet to
check which rule set produced it. Flattening both onto one sheet makes the
second invisible and the first cluttered.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from amrss.reports.build import ReportContent

HEADER_FILL = PatternFill("solid", fgColor="1F7A4C")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True)
MUTED_FONT = Font(color="5C6B64", size=9)
THIN = Side(style="thin", color="DCE7E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

#: Withheld cells are shaded neutrally, never in the S/I/R palette. A grey cell
#: reads as "no value here"; a red one would read as a resistance finding.
WITHHELD_FILL = PatternFill("solid", fgColor="EDEFEE")


def render(content: ReportContent) -> bytes:
    workbook = Workbook()
    active = workbook.active
    assert active is not None  # a new Workbook always has one sheet
    _antibiogram_sheet(active, content)
    _coverage_sheet(workbook.create_sheet("Coverage and quality"), content)
    if content.signals:
        _signals_sheet(workbook.create_sheet("Signals"), content)
    _methodology_sheet(workbook.create_sheet("Methodology"), content)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _antibiogram_sheet(sheet: Worksheet, content: ReportContent) -> None:
    sheet.title = "Antibiogram"
    sheet["A1"] = content.title
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = f"{content.scope_label} · {content.period_start} to {content.period_end}"
    sheet["A3"] = (
        f"Generated {content.generated_at:%d %b %Y %H:%M} UTC · "
        f"{content.eligible_isolate_count:,} antibiogram-eligible isolates of "
        f"{content.raw_isolate_count:,}"
    )
    sheet["A3"].font = MUTED_FONT

    header_row = 5
    headers = ["Organism", "Kingdom", "Isolates", *content.antibiotics]
    for column, label in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for offset, row in enumerate(content.rows, start=1):
        line = header_row + offset
        sheet.cell(row=line, column=1, value=row.organism).border = BORDER
        sheet.cell(row=line, column=2, value=row.kingdom).border = BORDER
        sheet.cell(row=line, column=3, value=row.isolate_count).border = BORDER
        for index, cell_content in enumerate(row.cells):
            cell = sheet.cell(row=line, column=4 + index)
            # Percentage and n in one cell, never a bare percentage (SDD 11.3).
            cell.value = (
                f"{cell_content.display} {cell_content.n_display}".strip()
                if cell_content.reportable
                else cell_content.display
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            if not cell_content.reportable:
                cell.fill = WITHHELD_FILL

    notes_row = header_row + len(content.rows) + 2
    for offset, note in enumerate(content.notes):
        cell = sheet.cell(row=notes_row + offset, column=1, value=note)
        cell.font = MUTED_FONT
    framing = sheet.cell(
        row=notes_row + len(content.notes) + 1, column=1, value=content.clinical_framing
    )
    framing.font = Font(italic=True, size=9)

    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 10
    for index in range(len(content.antibiotics)):
        sheet.column_dimensions[get_column_letter(4 + index)].width = 16
    # The organism column stays visible while scrolling a fifteen-agent panel.
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=4)


def _coverage_sheet(sheet: Worksheet, content: ReportContent) -> None:
    """Coverage sits on its own sheet, not as a footnote.

    A reader who does not know that one laboratory in seven was excluded cannot
    judge the table, and a footnote at the bottom of a wide sheet is a footnote
    nobody scrolls to.
    """
    rows = [
        ("Scope", content.scope_label),
        ("Period", f"{content.period_start} to {content.period_end}"),
        ("Data last updated", content.data_last_updated or "Not recorded"),
        (
            "Facilities contributing",
            f"{content.facilities_contributing} of {content.facilities_expected}",
        ),
        (
            "Completeness",
            f"{content.completeness_percent:.0f}%" if content.completeness_percent else "—",
        ),
        ("Facilities excluded on quality grounds", content.facilities_excluded_by_quality),
        ("Isolates excluded on quality grounds", content.isolates_excluded_by_quality),
        ("Isolates in period", content.raw_isolate_count),
        ("Antibiogram-eligible isolates", content.eligible_isolate_count),
        ("Measurements awaiting interpretation", content.pending_interpretation_count),
        ("Reporting threshold", f"n ≥ {content.minimum_isolates}"),
        ("Suppression threshold", f"n < {content.small_cell_threshold} withheld"),
    ]
    sheet["A1"] = "Coverage and quality"
    sheet["A1"].font = TITLE_FONT
    for offset, (label, value) in enumerate(rows, start=3):
        sheet.cell(row=offset, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=offset, column=2, value=value)
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 40


def _signals_sheet(sheet: Worksheet, content: ReportContent) -> None:
    sheet["A1"] = "Emerging resistance signals"
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = (
        "Signals are screening prompts for expert review, not confirmed outbreaks, and are "
        "kept apart from the antibiogram because they do not meet the reporting threshold."
    )
    sheet["A2"].font = MUTED_FONT
    for offset, line in enumerate(content.signals, start=4):
        sheet.cell(row=offset, column=1, value=line)
    sheet.column_dimensions["A"].width = 110


def _methodology_sheet(sheet: Worksheet, content: ReportContent) -> None:
    """The rules that produced every number in the workbook.

    Included in the file rather than linked, because a report circulated by
    email travels away from the system that generated it, and a figure whose
    thresholds cannot be recovered is not auditable.
    """
    sheet["A1"] = "Methodology"
    sheet["A1"].font = TITLE_FONT

    row = 3
    for key, value in content.methodology.items():
        if isinstance(value, list):
            sheet.cell(row=row, column=1, value=key).font = Font(bold=True)
            row += 1
            for entry in value:
                sheet.cell(row=row, column=2, value=str(entry))
                row += 1
        elif isinstance(value, dict):
            sheet.cell(row=row, column=1, value=key).font = Font(bold=True)
            row += 1
            for inner_key, inner_value in value.items():
                sheet.cell(row=row, column=2, value=str(inner_key))
                sheet.cell(row=row, column=3, value=str(inner_value))
                row += 1
        else:
            sheet.cell(row=row, column=1, value=key).font = Font(bold=True)
            sheet.cell(row=row, column=2, value=str(value))
            row += 1

    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 46
    sheet.column_dimensions["C"].width = 60
