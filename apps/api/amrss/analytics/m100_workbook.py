"""Converting a laboratory's CLSI M100 workbook into importable breakpoint rows.

A laboratory rarely holds M100 as a tidy CSV. It holds the licensed document,
and whoever extracts it — by hand or by a PDF text parser — produces a workbook
that is *mostly* a breakpoint table and partly the page furniture that surrounded
it: footnote fragments promoted to agent names, a table legend caught in the disk
content column, a row label split across two lines.

This module reads that workbook and produces rows in the shape of
``clsi_m100.template.csv``. It exists so that the strict importer in
``breakpoint_import`` never has to soften: everything it emits parses cleanly,
and everything it cannot vouch for is dropped and *named* in a report the
operator reads before the table becomes clinical methodology.

Three rules govern every decision here, and none of them may be relaxed for
convenience:

1. **A value that does not parse exactly is never guessed at.** ``≤8`` is a
   susceptible bound; a bare ``8`` in the susceptible column is a broken
   extraction, not an inference to make.
2. **A criterion narrower than the group it sits under is dropped, not
   widened.** The oxacillin row scoped to *S. epidermidis* is not a
   *Staphylococcus* spp. criterion, and applying it to *S. aureus* would call
   resistant isolates susceptible.
3. **Conflicting criteria are reported, never resolved.** Where the extraction
   collapsed two sub-tables into one — Salmonella and Shigella, meningitis and
   non-meningitis — the two candidates are dropped together. Picking one would
   be a coin toss on a threshold that reaches a patient.

Nothing here ships breakpoint values; it is a reader for a file the laboratory
supplies. See packages/clsi/README.md.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import IO

from openpyxl import load_workbook

#: Sheet names the extraction may use, in order of preference. The combined
#: sheet carries both methods per agent and is what M100 itself prints.
COMBINED_SHEETS = ("Combined (Zone + MIC)", "Combined", "Breakpoints")

#: Dashes CLSI uses for "no criterion defined". Distinct from a missing cell:
#: both mean absent here, but only the dash means the table looked and found none.
_ABSENT = {"", "-", "–", "—", "‒", "−", "n/a", "na", "none"}

#: Footnote markers the extraction leaves attached to values. ``^`` flags agents
#: that concentrate in urine; it qualifies the category, not the number. Only
#: symbols are stripped — a trailing letter is left alone, because ``10 units``
#: and ``10 µg`` are disk contents, not footnotes, and a strip that ate them
#: would turn a readable criterion into a silent drop.
_TRAILING_MARKERS = re.compile(r"[\^\*†‡§¶]+\s*$")

#: The micro sign and Greek mu are visually identical and both appear in
#: extractions of the same document.
_MU = str.maketrans({"μ": "µ"})

#: Qualifiers that appear beside an agent name and change what the criterion
#: applies to. Mapped onto the template's site/route columns so the engine can
#: pick the right one per isolate rather than treating them as duplicates.
_SITE_QUALIFIERS = {
    "meningitis": "meningitis",
    "nonmeningitis": "non_meningitis",
    "non-meningitis": "non_meningitis",
    "uncomplicated uti": "uti_uncomplicated",
    "uti": "uti_uncomplicated",
}
_ROUTE_QUALIFIERS = {
    "oral": "oral",
    "oral penicillin v": "oral",
    "parenteral": "iv",
    "iv": "iv",
    "intravenous": "iv",
}

#: A qualifier naming a *subset* of the row's organism group — "S. epidermidis",
#: "S. aureus, including". The criterion is narrower than the group it was
#: filed under and the extraction lost which sub-table it came from, so it is
#: dropped (rule 2). ``All staphylococci`` is the opposite case: it restates the
#: group, so it is kept.
_TAXONOMIC_HINT = re.compile(
    r"\bspp\.|\b[A-Z]\.\s*[a-z]|staphylococci|streptococci|enterococci|"
    r"MRSA|MSSA|SOSA|including",
    re.IGNORECASE,
)
_CONFIRMS_GROUP = re.compile(r"^all\b", re.IGNORECASE)

#: Reporting-tier and status markers: "(U)a", "(U, Inv.)a", "(Inv.)". They may
#: sit mid-label where a neighbouring column bled into it.
_STATUS_MARKER = re.compile(r"\s*\((?:U|Inv\.|U,\s*Inv\.)\)\s*[a-z]?", re.IGNORECASE)

#: A plausible disk content: one or two concentrations and a unit.
_DISK_CONTENT = re.compile(r"\d+(?:\.\d+)?(?:/\d+(?:\.\d+)?)?\s*(?:µg|ug|mcg|units|mg)")

#: A qualifier that is really a fragment of a footnote or of the disk-content
#: column, not a scope. A bare number is column bleed; a comparator or the word
#: MIC is footnote prose. Either way the label is damaged beyond mechanical
#: repair and the row is dropped rather than read.
_FRAGMENT_HINT = re.compile(r"^\d|[≤≥<>]|\bMICs?\b", re.IGNORECASE)

TEMPLATE_COLUMNS = (
    "organism_group",
    "agent_code",
    "method",
    "disk_content",
    "site",
    "route",
    "standard",
    "table_reference",
    "tier",
    "mic_susceptible_max",
    "mic_sdd_min",
    "mic_sdd_max",
    "mic_intermediate_min",
    "mic_intermediate_max",
    "mic_resistant_min",
    "disk_susceptible_min",
    "disk_sdd_min",
    "disk_sdd_max",
    "disk_intermediate_min",
    "disk_intermediate_max",
    "disk_resistant_max",
    "dosage_note",
    "comment",
)


class WorkbookFormatError(ValueError):
    """The file is not a breakpoint workbook this reader understands."""


@dataclass(frozen=True)
class Dropped:
    """One source row that produced no criterion, and why.

    Carries the sheet row number so the operator can put the workbook and the
    printed table side by side. A drop reported without its line number is a
    drop nobody checks.
    """

    sheet_row: int
    organism_group: str
    agent_label: str
    reason: str

    def __str__(self) -> str:
        return f"row {self.sheet_row}: {self.agent_label!r} ({self.organism_group}) — {self.reason}"


@dataclass(frozen=True)
class Repair:
    """A source row whose agent was identified by repairing a damaged label.

    Every repair is surfaced for review. The rules are mechanical, but a
    mechanical rule applied to a lossy extraction still deserves a human eye
    before it interprets a patient's isolate.
    """

    sheet_row: int
    agent_label: str
    agent_code: str
    rule: str

    def __str__(self) -> str:
        return f"row {self.sheet_row}: {self.agent_label!r} read as {self.agent_code} ({self.rule})"


@dataclass
class WorkbookConversion:
    rows: list[dict[str, object]] = field(default_factory=list)
    dropped: list[Dropped] = field(default_factory=list)
    repairs: list[Repair] = field(default_factory=list)

    @property
    def organism_groups(self) -> list[str]:
        return sorted({str(r["organism_group"]) for r in self.rows})

    @property
    def agent_codes(self) -> list[str]:
        return sorted({str(r["agent_code"]) for r in self.rows})


# --------------------------------------------------------------------------
# Value parsing
#
# Every parser returns None for an absent value and raises for a present one it
# cannot read exactly. That asymmetry is the whole safety property: a blank
# means "M100 defines nothing here", and anything else must be understood or
# refused.
# --------------------------------------------------------------------------


class _UnreadableValueError(ValueError):
    pass


def _text(value: object) -> str:
    return "" if value is None else str(value).translate(_MU).strip()


def _strip_markers(raw: str) -> str:
    return _TRAILING_MARKERS.sub("", raw.strip()).strip()


def _is_absent(raw: str) -> bool:
    return _strip_markers(raw).lower() in _ABSENT


def _component(token: str) -> str:
    """The first component of a combination value.

    CLSI writes combination agents as ``≤8/4``: the breakpoint is the active
    agent's concentration, with the partner fixed at the second figure. The
    partner is not a threshold and must not be read as one.
    """
    return token.split("/", 1)[0].strip()


def _decimal(token: str) -> Decimal:
    try:
        value = Decimal(_component(token))
    except InvalidOperation as exc:
        raise _UnreadableValueError(f"{token!r} is not a concentration") from exc
    if value <= 0:
        raise _UnreadableValueError(f"{token!r} is not a positive concentration")
    return value


def _integer(token: str) -> int:
    text = _component(token)
    if not re.fullmatch(r"\d+", text):
        raise _UnreadableValueError(f"{token!r} is not a zone diameter")
    return int(text)


def _bounded(raw: str, prefix: str, parse):
    """Read a one-sided bound such as ``≤8`` or ``≥17``.

    The comparator is required. A bare number in a bound column means the
    extraction lost the symbol, and without it there is no way to tell a
    susceptible ceiling from a resistant floor.
    """
    text = _strip_markers(raw)
    if not text.startswith(prefix):
        raise _UnreadableValueError(f"{raw!r} is missing its {prefix} comparator")
    return parse(text[len(prefix) :])


def _range(raw: str, parse):
    """Read an intermediate or SDD range: ``4``, ``4-8``, or ``1/19-2/38``."""
    text = _strip_markers(raw)
    if text.startswith("≤"):
        # An SDD ceiling with no floor, e.g. fluconazole SDD ≤4.
        return None, parse(text[1:])
    parts = [p for p in re.split(r"(?<![/\d])-|(?<=\d)-(?=\d)", text) if p.strip()]
    if len(parts) == 1:
        # A single dilution is a range of one.
        value = parse(parts[0])
        return value, value
    if len(parts) == 2:
        low, high = parse(parts[0]), parse(parts[1])
        if low > high:
            raise _UnreadableValueError(f"{raw!r} runs backwards")
        return low, high
    raise _UnreadableValueError(f"{raw!r} is not a range")


# --------------------------------------------------------------------------
# Agent identification
# --------------------------------------------------------------------------


def _normalise_label(label: str) -> str:
    text = label.replace("‑", "-").replace("’", "'").strip()
    text = _STATUS_MARKER.sub(" ", text)
    text = text.replace("*", " ")
    return re.sub(r"\s+", " ", text).strip()


@dataclass(frozen=True)
class _AgentMatch:
    code: str
    qualifier: str | None
    rule: str | None  # None where the label matched exactly


def resolve_agent(label: str, agents: dict[str, str]) -> _AgentMatch | None:
    """Identify the dictionary agent a workbook label refers to.

    ``agents`` maps a lower-cased agent name (and any CLSI alias the dictionary
    records) to its code — supplied by the caller rather than hardcoded, so a
    laboratory that adds an agent does not need a code change (ADR-0002).

    Three rules, tried in order, each mechanical:

    1. the label *is* an agent name;
    2. the label is an agent name followed by a qualifier the extraction merged
       in from a neighbouring column — ``Cefepime (meningitis)``,
       ``Penicillin All staphylococci``;
    3. the label is a truncated agent name — ``Trimethoprim-`` — and exactly one
       agent name begins with it. Ambiguity here resolves to no match, never to
       the first candidate.

    Anything else is a footnote fragment, a table legend, or an agent the
    dictionary does not carry, and returns None.
    """
    text = _normalise_label(label)
    if not text:
        return None

    key = text.lower()
    if key in agents:
        return _AgentMatch(agents[key], None, None)

    parenthetical = re.fullmatch(r"(.+?)\s*\((.+)\)", text)
    if parenthetical:
        head, inside = parenthetical.group(1).strip(), parenthetical.group(2).strip()
        if head.lower() in agents:
            return _AgentMatch(agents[head.lower()], inside, "qualifier in parentheses")

    # Longest agent name first, so "Amoxicillin-clavulanate" is preferred over
    # "Amoxicillin" when the label carries both.
    for name in sorted(agents, key=len, reverse=True):
        if key.startswith(name + " "):
            return _AgentMatch(agents[name], text[len(name) :].strip(), "trailing qualifier")

    if text.endswith("-"):
        candidates = {code for name, code in agents.items() if name.startswith(key)}
        if len(candidates) == 1:
            return _AgentMatch(next(iter(candidates)), None, "truncated agent name")

    return None


def _classify_qualifier(qualifier: str) -> tuple[str | None, str | None, str | None]:
    """Split a label qualifier into (site, route, drop-reason).

    M100 prints alternatives on consecutive lines — "Cefotaxime or" above
    "ceftriaxone" — so a leading "or" is line furniture and is discarded; the
    agent on this line is the one this row's numbers belong to.
    """
    text = qualifier.strip().strip(",").lower()
    text = re.sub(r"^or\s+", "", text).strip()
    if not text:
        return None, None, None
    if text in _SITE_QUALIFIERS:
        return _SITE_QUALIFIERS[text], None, None
    if text in _ROUTE_QUALIFIERS:
        return None, _ROUTE_QUALIFIERS[text], None
    if _FRAGMENT_HINT.search(text):
        return (
            None,
            None,
            f"the label carries {qualifier!r}, which is a fragment of a "
            "neighbouring column or footnote rather than a scope",
        )
    if _CONFIRMS_GROUP.match(text):
        return None, None, None  # "All staphylococci" restates the row's group
    if _TAXONOMIC_HINT.search(text):
        return (
            None,
            None,
            f"criterion is scoped to {qualifier!r}, which is narrower than its "
            "organism group; the extraction lost which sub-table it came from",
        )
    return None, None, None


# --------------------------------------------------------------------------
# Conversion
# --------------------------------------------------------------------------

_COMBINED_HEADER = (
    "table",
    "organism",
    "drug class",
    "antimicrobial agent",
    "disk content",
)


def _locate_data_start(sheet) -> int:
    """Find the first row of data.

    The combined sheet carries a two-line header: a merged method band above the
    S/SDD/I/R quartets. Locating it by content, and detecting the sub-header by
    its empty leading cells rather than by a fixed offset, means a workbook
    exported with a title row above the table still reads — and one without a
    sub-header does not lose its first criterion.
    """
    rows = list(sheet.iter_rows(min_row=1, max_row=12, values_only=True))
    for index, row in enumerate(rows, start=1):
        if tuple(_text(c).lower() for c in row[:5]) != _COMBINED_HEADER:
            continue
        following = rows[index] if index < len(rows) else ()
        is_subheader = not any(_text(c) for c in following[:5])
        return index + (2 if is_subheader else 1)
    raise WorkbookFormatError(
        "no breakpoint header row found. Expected a sheet whose first columns are "
        "Table, Organism, Drug Class, Antimicrobial Agent, Disk Content."
    )


def convert_workbook(source: str | IO[bytes], *, agents: dict[str, str]) -> WorkbookConversion:
    """Read an M100 workbook into template rows, a drop list and a repair list.

    ``agents`` maps lower-cased agent names and CLSI aliases to dictionary
    codes. An agent the dictionary does not carry produces a drop, not a row:
    a breakpoint whose agent cannot be joined to a result is inert, and a silent
    one would let a laboratory believe a table is complete when it is not.
    """
    workbook = load_workbook(source, read_only=True, data_only=True)
    name = next((s for s in COMBINED_SHEETS if s in workbook.sheetnames), None)
    if name is None:
        raise WorkbookFormatError(
            "no combined breakpoint sheet found. Expected one of: "
            + ", ".join(COMBINED_SHEETS)
            + f". This file has: {', '.join(workbook.sheetnames)}"
        )
    sheet = workbook[name]
    first_data_row = _locate_data_start(sheet)

    conversion = WorkbookConversion()
    #: (organism_group, agent_code, method, site, route) -> row indices, so a
    #: collision can withdraw *both* candidates rather than the later one.
    seen: dict[tuple, list[int]] = {}

    for offset, raw in enumerate(
        sheet.iter_rows(min_row=first_data_row, values_only=True), start=first_data_row
    ):
        _convert_row(raw, offset, agents, conversion, seen)

    _withdraw_conflicts(conversion, seen)
    workbook.close()
    return conversion


def _convert_row(
    raw: tuple,
    sheet_row: int,
    agents: dict[str, str],
    conversion: WorkbookConversion,
    seen: dict[tuple, list[int]],
) -> None:
    cells = [_text(c) for c in raw] + [""] * 14
    table_ref, organism, _drug_class, agent_label, disk_content = cells[:5]
    zone = cells[5:9]
    mic = cells[9:13]
    comments = cells[13]

    if not organism or not agent_label:
        return

    def drop(reason: str) -> None:
        conversion.dropped.append(Dropped(sheet_row, organism, agent_label, reason))

    match = resolve_agent(agent_label, agents)
    if match is None:
        drop("no agent in the dictionary matches this label")
        return

    site, route, refusal = _classify_qualifier(match.qualifier or "")
    if refusal:
        drop(refusal)
        return
    if match.rule:
        conversion.repairs.append(Repair(sheet_row, agent_label, match.code, match.rule))

    base: dict[str, object] = {
        "organism_group": re.sub(r"\s+", " ", organism),
        "agent_code": match.code,
        "site": site,
        "route": route,
        "standard": "M100",
        "table_reference": f"Table {table_ref}" if table_ref else None,
        "tier": None,
        "dosage_note": None,
        "comment": (comments or None),
    }

    produced = 0
    for method, values, parse in (("MIC", mic, _decimal), ("DISK", zone, _integer)):
        try:
            bounds = _read_categories(values, parse)
        except _UnreadableValueError as exc:
            drop(f"{method} column unreadable: {exc}")
            continue
        if bounds is None:
            continue
        if method == "DISK":
            content = _strip_markers(disk_content)
            if not _DISK_CONTENT.fullmatch(content):
                drop(
                    f"disk criteria present but the disk content {disk_content!r} is "
                    "not a concentration; a zone diameter without its disk content "
                    "cannot be interpreted"
                )
                continue
            row = base | {"method": "DISK", "disk_content": content} | _disk_columns(bounds)
        else:
            row = base | {"method": "MIC", "disk_content": None} | _mic_columns(bounds)

        seen.setdefault((row["organism_group"], row["agent_code"], method, site, route), []).append(
            len(conversion.rows)
        )
        conversion.rows.append(row)
        produced += 1

    if produced == 0 and not any(
        d.sheet_row == sheet_row for d in conversion.dropped
    ):  # pragma: no branch
        drop("no breakpoint values on this row")


def _read_categories(values: list[str], parse):
    """Read the S / SDD / I / R quartet for one method, or None if all absent."""
    s_raw, sdd_raw, i_raw, r_raw = values
    if all(_is_absent(v) for v in (s_raw, sdd_raw, i_raw, r_raw)):
        return None

    comparators = ("≤", "≥") if parse is _decimal else ("≥", "≤")
    susceptible = None if _is_absent(s_raw) else _bounded(s_raw, comparators[0], parse)
    resistant = None if _is_absent(r_raw) else _bounded(r_raw, comparators[1], parse)
    sdd = (None, None) if _is_absent(sdd_raw) else _range(sdd_raw, parse)
    intermediate = (None, None) if _is_absent(i_raw) else _range(i_raw, parse)

    if susceptible is None and resistant is None:
        raise _UnreadableValueError("neither a susceptible nor a resistant bound")
    return susceptible, sdd, intermediate, resistant


def _mic_columns(bounds) -> dict[str, object]:
    susceptible, sdd, intermediate, resistant = bounds
    return {
        "mic_susceptible_max": susceptible,
        "mic_sdd_min": sdd[0],
        "mic_sdd_max": sdd[1],
        "mic_intermediate_min": intermediate[0],
        "mic_intermediate_max": intermediate[1],
        "mic_resistant_min": resistant,
    }


def _disk_columns(bounds) -> dict[str, object]:
    susceptible, sdd, intermediate, resistant = bounds
    return {
        "disk_susceptible_min": susceptible,
        "disk_sdd_min": sdd[0],
        "disk_sdd_max": sdd[1],
        "disk_intermediate_min": intermediate[0],
        "disk_intermediate_max": intermediate[1],
        "disk_resistant_max": resistant,
    }


def _withdraw_conflicts(conversion: WorkbookConversion, seen: dict[tuple, list[int]]) -> None:
    """Drop every member of a colliding set of criteria.

    Where two rows claim the same organism group, agent, method and qualifiers
    with different numbers, the extraction has flattened two sub-tables into one
    — Salmonella beside Shigella, meningitis beside non-meningitis. Keeping
    either would be a guess about which patients it applies to, and keeping the
    last one silently would make the answer depend on row order.
    """
    doomed: set[int] = set()
    for key, indices in seen.items():
        if len(indices) < 2:
            continue
        group, code, method, *_ = key
        doomed.update(indices)
        conversion.dropped.append(
            Dropped(
                0,
                str(group),
                str(code),
                f"{len(indices)} conflicting {method} criteria for this organism group "
                "and agent. M100 tabulates these separately; the extraction merged "
                "them. Resolve by hand against the printed table.",
            )
        )
    if doomed:
        conversion.rows = [r for i, r in enumerate(conversion.rows) if i not in doomed]


def to_template_csv(rows: list[dict[str, object]]) -> str:
    """Render converted rows in the shape of ``clsi_m100.template.csv``.

    The CSV is the reviewable artefact. An operator can diff it against the
    printed tables, and it is what the strict importer consumes — the workbook
    reader gets no privileged path into clinical methodology.
    """
    import csv
    import io

    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=list(TEMPLATE_COLUMNS), extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({c: ("" if row.get(c) is None else row[c]) for c in TEMPLATE_COLUMNS})
    return buffer.getvalue()
